import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os, threading, math

# ============================================================
# CONFIGURAÇÕES — colunas
# ============================================================
CG_CODIGO       = "Codigo"
CG_NOME         = "Nome"
CG_INSTALACAO   = "Instalacao"
CG_DATA_ATIVO   = "Data Ativo"
CG_DATA_INJECAO = "Data Injecao"
CG_FORNECEDORA  = "Fornecedora"

REC_CODIGO      = "Codigo Cliente"
REC_INSTALACAO  = "Instalacao"
REC_DATA_REF    = "Data Referencia"
REC_FORNECEDORA = "Fornecedora"

DIAS_CARENCIA   = 90

# ============================================================
# PALETA iGreen dark
# ============================================================
BG         = "#0e1117"
SURFACE    = "#161b24"
SURFACE2   = "#1c2333"
SURFACE3   = "#222d3a"
BORDER     = "#263040"
ACCENT     = "#22c55e"
ACCENT_DIM = "#16a34a"
RED        = "#ef4444"
AMBER      = "#f59e0b"
BLUE       = "#3b82f6"
TEXT       = "#f1f5f9"
TEXT2      = "#94a3b8"
TEXT3      = "#64748b"
LOG_BG     = "#0D1F15"
LOG_FG     = "#4ade80"

# ============================================================
# UTILITÁRIOS
# ============================================================
def parse_data(val):
    if pd.isna(val) or str(val).strip() in ("", "nan", "None"): return None
    if isinstance(val, datetime): return val
    if hasattr(val, 'to_pydatetime'): return val.to_pydatetime()
    s = str(val).strip()
    for fmt in ("%d/%m/%Y","%Y-%m-%d","%d-%m-%Y","%m/%d/%Y","%Y/%m/%d"):
        try: return datetime.strptime(s, fmt)
        except: pass
    try:
        n = float(s)
        if n > 40000: return datetime(1899,12,30)+timedelta(days=int(n))
    except: pass
    return None

def normalizar(s):
    return "" if pd.isna(s) else str(s).strip().upper()

def fmt_data(dt):
    return "" if dt is None else dt.strftime("%d/%m/%Y")

def gerar_meses(di, df):
    meses, dt = [], datetime(di.year, di.month, 1)
    fim = datetime(df.year, df.month, 1)
    while dt <= fim:
        meses.append(dt.strftime("%m/%Y"))
        dt = datetime(dt.year+1,1,1) if dt.month==12 else datetime(dt.year, dt.month+1, 1)
    return meses

def extrair_mes(val):
    if pd.isna(val) or str(val).strip() in ("","nan","None"): return None
    s = str(val).strip()
    if len(s)==7 and s[2]=="/": return s
    dt = parse_data(val)
    return dt.strftime("%m/%Y") if dt else None

# ============================================================
# LÓGICA
# ============================================================
def processar(df_cl, df_rec, forn_filtro, log_fn):
    hoje = datetime.today().replace(hour=0,minute=0,second=0,microsecond=0)
    df_cl.columns  = df_cl.columns.str.strip()
    df_rec.columns = df_rec.columns.str.strip()

    if forn_filtro.strip():
        fu = forn_filtro.strip().upper()
        if CG_FORNECEDORA in df_cl.columns:
            df_cl = df_cl[df_cl[CG_FORNECEDORA].astype(str).str.upper().str.contains(fu)].copy()
        if REC_FORNECEDORA in df_rec.columns:
            df_rec = df_rec[df_rec[REC_FORNECEDORA].astype(str).str.upper().str.contains(fu)].copy()

    log_fn(f"Clientes: {len(df_cl):,}  |  Recebíveis: {len(df_rec):,}")

    if REC_INSTALACAO in df_rec.columns:
        df_rec["__inst"] = df_rec[REC_INSTALACAO].apply(normalizar)
    else:
        df_rec["__inst"] = ""

    if REC_DATA_REF in df_rec.columns:
        df_rec["__mes"] = df_rec[REC_DATA_REF].apply(extrair_mes)
    else:
        df_rec["__mes"] = None

    idx_boletos = df_rec.groupby("__inst")["__mes"].apply(
        lambda x: [m for m in x if m]
    ).to_dict()

    rows = []
    log_fn("Calculando diferenças por instalação...")

    for _, r in df_cl.iterrows():
        inst  = normalizar(r.get(CG_INSTALACAO,""))
        cod   = str(r.get(CG_CODIGO,"")).strip()
        nome  = str(r.get(CG_NOME,"")).strip()
        forn  = str(r.get(CG_FORNECEDORA,"")).strip()
        da    = parse_data(r.get(CG_DATA_ATIVO) or r.get("Data Ativacao Original"))
        di    = parse_data(r.get(CG_DATA_INJECAO))
        if di is None and da: di = da + timedelta(days=DIAS_CARENCIA)
        exist = idx_boletos.get(inst, [])
        be    = len(exist)

        if da is None:
            rows.append({"Instalacao":inst,"Codigo":cod,"Nome":nome,"Fornecedora":forn,
                         "Data Ativacao":"","Data 1a Injecao":"",
                         "Boletos Existentes":be,"Boletos Esperados":"",
                         "Diferenca":"","Meses Faltantes":"","Status":"SEM DATA ATIVACAO"})
            continue

        if di and di > hoje:
            rows.append({"Instalacao":inst,"Codigo":cod,"Nome":nome,"Fornecedora":forn,
                         "Data Ativacao":fmt_data(da),"Data 1a Injecao":fmt_data(di),
                         "Boletos Existentes":be,"Boletos Esperados":0,
                         "Diferenca":"","Meses Faltantes":"","Status":"AINDA NAO FATURA"})
            continue

        bexp  = max(0, math.floor((hoje - di).days / 30)) if di else 0
        diff  = be - bexp
        falt  = ""
        if diff < 0 and di:
            todos  = set(gerar_meses(di, hoje))
            falt   = ", ".join(sorted(todos - set(exist)))

        status = "POSITIVO" if diff > 0 else ("NEGATIVO" if diff < 0 else "ZERO")
        rows.append({"Instalacao":inst,"Codigo":cod,"Nome":nome,"Fornecedora":forn,
                     "Data Ativacao":fmt_data(da),"Data 1a Injecao":fmt_data(di),
                     "Boletos Existentes":be,"Boletos Esperados":bexp,
                     "Diferenca":diff,"Meses Faltantes":falt,"Status":status})

    return pd.DataFrame(rows)

# ============================================================
# EXCEL
# ============================================================
def formatar_excel(path):
    wb = load_workbook(path)
    # Cabecalho: fundo verde escuro iGreen, texto branco negrito
    hf  = PatternFill("solid", fgColor="166534")
    hft = Font(bold=True, color="FFFFFF", size=11)
    ha  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # Linhas alternadas: branco e cinza claro
    fp  = PatternFill("solid", fgColor="FFFFFF")
    fi  = PatternFill("solid", fgColor="F3F4F6")
    ft_normal = Font(color="111827", size=10)
    for nome in wb.sheetnames:
        ws = wb[nome]
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 28
        for c in ws[1]:
            c.fill = hf; c.font = hft; c.alignment = ha
        for i, row in enumerate(ws.iter_rows(min_row=2), 2):
            f = fp if i % 2 == 0 else fi
            for c in row:
                c.fill = f
                c.font = ft_normal
                c.alignment = Alignment(vertical="center")
        for col in ws.iter_cols(max_row=min(200, ws.max_row)):
            ml = 0; cl = get_column_letter(col[0].column)
            for c in col:
                if c.value: ml = max(ml, len(str(c.value)))
            ws.column_dimensions[cl].width = min(ml + 4, 50)
    wb.save(path)

def salvar_excel(df, path):
    neg   = df[df["Status"]=="NEGATIVO"]
    pos   = df[df["Status"]=="POSITIVO"]
    zero  = df[df["Status"]=="ZERO"]
    ainda = df[df["Status"]=="AINDA NAO FATURA"]
    sem   = df[df["Status"]=="SEM DATA ATIVACAO"]
    res   = pd.DataFrame([
        {"Status":"🔴 NEGATIVO — Boletos Faltantes",     "Qtd":len(neg)},
        {"Status":"🟡 POSITIVO — Possível Duplicado",    "Qtd":len(pos)},
        {"Status":"🟢 ZERO — OK",                       "Qtd":len(zero)},
        {"Status":"⚪ AINDA NÃO FATURA (<90 dias)",     "Qtd":len(ainda)},
        {"Status":"❓ SEM DATA DE ATIVAÇÃO",             "Qtd":len(sem)},
        {"Status":"TOTAL",                               "Qtd":len(df)},
    ])
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        res.to_excel(w,   sheet_name="RESUMO",          index=False)
        neg.to_excel(w,   sheet_name="NEGATIVOS",       index=False)
        pos.to_excel(w,   sheet_name="POSITIVOS",       index=False)
        zero.to_excel(w,  sheet_name="ZERO",            index=False)
        ainda.to_excel(w, sheet_name="AINDA NAO FATURA",index=False)
        if not sem.empty:
            sem.to_excel(w, sheet_name="SEM DATA",      index=False)
    formatar_excel(path)

# ============================================================
# INTERFACE — paleta dark iGreen
# ============================================================
# ============================================================
# PIXEL ART — tela de carregamento
# ============================================================
PIXEL_SIZE = 10

# Pixel art de um raio (símbolo iGreen energia) — grid 9x14
RAIO = [
    "....XX...",
    "...XXX...",
    "..XXXX...",
    ".XXXXX...",
    "XXXXXX...",
    ".XXXXXX..",
    "..XXXXXX.",
    "...XXXXX.",
    "....XXXX.",
    ".....XXX.",
    "......XX.",
    ".......X.",
    "......X..",
    ".....X...",
]

class LoadingScreen(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("")
        self.configure(bg=BG)
        self.resizable(False, False)
        self.overrideredirect(True)  # sem barra de título

        W, H = 320, 260
        px = parent.winfo_x() + (parent.winfo_width()  - W) // 2
        py = parent.winfo_y() + (parent.winfo_height() - H) // 2
        self.geometry(f"{W}x{H}+{px}+{py}")

        # Borda
        tk.Frame(self, bg=BORDER, bd=1, relief="solid").place(relwidth=1, relheight=1)
        inner = tk.Frame(self, bg=BG)
        inner.place(x=1, y=1, width=W-2, height=H-2)

        # Canvas pixel art
        cols  = len(RAIO[0])
        rows  = len(RAIO)
        cw    = cols * PIXEL_SIZE
        ch    = rows * PIXEL_SIZE
        self.canvas = tk.Canvas(inner, width=cw, height=ch,
                                bg=BG, highlightthickness=0)
        self.canvas.place(x=(W-2-cw)//2, y=24)

        self._pixels = []
        for r, linha in enumerate(RAIO):
            row_pix = []
            for c, ch_ in enumerate(linha):
                x0 = c * PIXEL_SIZE
                y0 = r * PIXEL_SIZE
                cor = ACCENT if ch_ == "X" else BG
                px_id = self.canvas.create_rectangle(
                    x0, y0, x0+PIXEL_SIZE-1, y0+PIXEL_SIZE-1,
                    fill=cor, outline="")
                row_pix.append((px_id, ch_ == "X"))
            self._pixels.append(row_pix)

        # Texto
        self.msg = tk.Label(inner, text="Processando análise...",
                            bg=BG, fg=TEXT2, font=("Segoe UI", 10))
        self.msg.place(x=0, y=H-80, width=W-2)

        self.sub = tk.Label(inner, text="Cruzando planilhas",
                            bg=BG, fg=TEXT3, font=("Segoe UI", 8))
        self.sub.place(x=0, y=H-58, width=W-2)

        # Barra de progresso animada
        self._bar_frame = tk.Frame(inner, bg=SURFACE2, height=3)
        self._bar_frame.place(x=24, y=H-36, width=W-50)
        self._bar = tk.Frame(self._bar_frame, bg=ACCENT, height=3)
        self._bar.place(x=0, y=0, width=0, height=3)
        self._bar_w = 0
        self._bar_dir = 1

        self._frame = 0
        self._animating = True
        self._animate()

    def _animate(self):
        if not self._animating:
            return
        self._frame += 1
        t = self._frame

        # Pulsa os pixels do raio
        for r, row in enumerate(self._pixels):
            for c, (pid, is_lit) in enumerate(row):
                if is_lit:
                    import math as m
                    wave = m.sin((t * 0.15) + (r + c) * 0.4)
                    bright = int(80 + 80 * wave)  # 0–160 offset do verde
                    g = min(255, 197 + bright // 3)
                    b = min(255, 94  + bright // 4)
                    cor = f"#22{g:02x}{b:02x}"
                    self.canvas.itemconfig(pid, fill=cor)

        # Barra bouncing
        bar_total = self._bar_frame.winfo_width() or 272
        speed = 4
        self._bar_w += speed * self._bar_dir
        if self._bar_w >= bar_total: self._bar_dir = -1
        if self._bar_w <= 0:         self._bar_dir =  1
        self._bar.place(x=0, y=0, width=max(0, self._bar_w), height=3)

        # Mensagem rotativa
        msgs = ["Cruzando planilhas", "Calculando boletos", "Identificando faltantes", "Gerando relatório"]
        self.sub.config(text=msgs[(t // 40) % len(msgs)] + "...")

        self.after(40, self._animate)

    def atualizar(self, texto):
        self.msg.config(text=texto)

    def fechar(self):
        self._animating = False
        self.destroy()


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("iGreen — Análise de Faturamento")
        self.root.configure(bg=BG)
        self.root.geometry("740x680")
        self.root.resizable(False, False)

        self.arq_cl  = tk.StringVar()
        self.arq_rec = tk.StringVar()
        self.forn    = tk.StringVar()

        self._ui()

    def _ui(self):
        # ── Topbar ──
        top = tk.Frame(self.root, bg=SURFACE, height=56)
        top.pack(fill="x"); top.pack_propagate(False)

        logo_box = tk.Frame(top, bg=ACCENT, width=32, height=32)
        logo_box.place(x=18, y=12)
        tk.Label(logo_box, text="iG", bg=ACCENT, fg="#000",
                 font=("Segoe UI", 12, "bold")).place(relx=.5, rely=.5, anchor="center")

        tk.Label(top, text="iGreen Energy", bg=SURFACE, fg=TEXT,
                 font=("Segoe UI", 13, "bold")).place(x=60, y=10)
        tk.Label(top, text="Análise de Faturamento de Boletos", bg=SURFACE, fg=TEXT3,
                 font=("Segoe UI", 9)).place(x=61, y=32)

        # Separador
        tk.Frame(self.root, bg=BORDER, height=1).pack(fill="x")

        # ── Corpo ──
        body = tk.Frame(self.root, bg=BG)
        body.pack(fill="both", expand=True, padx=24, pady=20)

        # Título seção
        self._sec_label(body, "Planilhas de Entrada")

        # Cards de upload
        self._upload_card(body, "📋  Clientes Green (Base Completa)", "BackOffice — obrigatório", self.arq_cl, 0)
        self._upload_card(body, "💰  Recebíveis",                     "BackOffice — obrigatório", self.arq_rec, 1)

        # Filtro fornecedora
        self._sec_label(body, "Filtro de Fornecedora", mt=14)
        forn_frame = tk.Frame(body, bg=SURFACE, bd=0, relief="flat",
                              highlightbackground=BORDER, highlightthickness=1)
        forn_frame.pack(fill="x", pady=(4,0))

        tk.Label(forn_frame, text="Fornecedora:", bg=SURFACE, fg=TEXT2,
                 font=("Segoe UI", 10)).pack(side="left", padx=(14,8), pady=10)

        self.forn_entry = tk.Entry(forn_frame, textvariable=self.forn,
                                   bg=SURFACE2, fg=TEXT, insertbackground=TEXT,
                                   relief="flat", font=("Segoe UI", 11), width=18,
                                   highlightbackground=BORDER, highlightthickness=1)
        self.forn_entry.pack(side="left", pady=8)

        tk.Label(forn_frame, text="GV  ·  EDP  ·  SUNNE  ·  BC  ·  POPSOL  ·  FIT   (vazio = todas)",
                 bg=SURFACE, fg=TEXT3, font=("Segoe UI", 9)).pack(side="left", padx=12)

        # Botão processar
        btn_frame = tk.Frame(self.root, bg=BG)
        btn_frame.pack(fill="x", padx=24, pady=(12,0))

        self.btn = tk.Button(
            btn_frame, text="▶   Processar Análise",
            command=self._iniciar,
            bg=ACCENT, fg="#000", relief="flat",
            font=("Segoe UI", 12, "bold"),
            cursor="hand2", padx=28, pady=11,
            activebackground=ACCENT_DIM, activeforeground="#000")
        self.btn.pack(side="right")

        # Status pill
        self.status_var = tk.StringVar(value="Aguardando arquivos...")
        tk.Label(btn_frame, textvariable=self.status_var,
                 bg=BG, fg=TEXT3, font=("Segoe UI", 9)).pack(side="left", pady=4)

        # Separador
        tk.Frame(self.root, bg=BORDER, height=1).pack(fill="x", padx=0, pady=(12,0))

        # Log
        log_hdr = tk.Frame(self.root, bg=SURFACE)
        log_hdr.pack(fill="x")
        tk.Label(log_hdr, text="  LOG DE EXECUÇÃO", bg=SURFACE, fg=TEXT3,
                 font=("Segoe UI", 8, "bold")).pack(side="left", pady=5)

        self.log = tk.Text(self.root, height=8, bg=LOG_BG, fg=LOG_FG,
                           font=("Consolas", 9), state="disabled", bd=0,
                           padx=12, pady=8, selectbackground=SURFACE2)
        self.log.pack(fill="x", padx=0, pady=0)

        # Footer
        tk.Frame(self.root, bg=BORDER, height=1).pack(fill="x")
        footer = tk.Frame(self.root, bg=SURFACE)
        footer.pack(fill="x")
        tk.Label(footer, text="iGreen Energy  ·  Gestão de Inadimplência  ·  N2",
                 bg=SURFACE, fg=TEXT3, font=("Segoe UI", 8)).pack(side="left", padx=12, pady=5)
        tk.Label(footer, text="Feito por Lucas Coutinho",
                 bg=SURFACE, fg=TEXT3, font=("Segoe UI", 8)).pack(side="right", padx=12, pady=5)

    def _sec_label(self, parent, txt, mt=0):
        tk.Label(parent, text=txt.upper(), bg=BG, fg=TEXT3,
                 font=("Segoe UI", 8, "bold")).pack(
            anchor="w", pady=(mt, 4))

    def _upload_card(self, parent, titulo, subtitulo, var, idx):
        card = tk.Frame(parent, bg=SURFACE,
                        highlightbackground=BORDER, highlightthickness=1)
        card.pack(fill="x", pady=(0,8))

        # Ícone estado
        self._dots = getattr(self, "_dots", {})
        dot = tk.Label(card, text="○", bg=SURFACE, fg=TEXT3,
                       font=("Segoe UI", 14))
        dot.pack(side="left", padx=(14,10), pady=10)
        self._dots[idx] = dot

        info = tk.Frame(card, bg=SURFACE)
        info.pack(side="left", fill="both", expand=True, pady=8)
        tk.Label(info, text=titulo, bg=SURFACE, fg=TEXT,
                 font=("Segoe UI", 10, "bold"), anchor="w").pack(anchor="w")
        lbl = tk.Label(info, text=subtitulo, bg=SURFACE, fg=TEXT3,
                       font=("Segoe UI", 9), anchor="w")
        lbl.pack(anchor="w")

        # Nome arquivo selecionado
        nome_var = tk.StringVar(value="Nenhum arquivo selecionado")
        nm = tk.Label(info, textvariable=nome_var, bg=SURFACE, fg=TEXT3,
                      font=("Segoe UI", 8), anchor="w")
        nm.pack(anchor="w")

        def selecionar():
            path = filedialog.askopenfilename(
                filetypes=[("Excel/CSV","*.xlsx *.xls *.csv"),("Todos","*.*")])
            if path:
                var.set(path)
                fname = os.path.basename(path)
                nome_var.set(f"✓  {fname}")
                nm.config(fg=ACCENT)
                dot.config(text="●", fg=ACCENT)
                self.status_var.set("Pronto para processar" if self.arq_cl.get() and self.arq_rec.get() else "Aguardando arquivos...")

        tk.Button(card, text="Selecionar",
                  command=selecionar,
                  bg=SURFACE2, fg=TEXT2, relief="flat",
                  font=("Segoe UI", 9, "bold"), cursor="hand2",
                  padx=14, pady=6,
                  activebackground=SURFACE3, activeforeground=TEXT,
                  highlightbackground=BORDER, highlightthickness=1).pack(
            side="right", padx=14, pady=10)

    def _log(self, msg):
        def _w():
            self.log.configure(state="normal")
            ts = datetime.now().strftime("%H:%M:%S")
            # Cor por tipo
            tag = "info"
            if "❌" in msg or "Erro" in msg: tag = "err"
            elif "✅" in msg: tag = "ok"
            elif "🔴" in msg: tag = "red"
            elif "🟡" in msg: tag = "amb"
            elif "🟢" in msg: tag = "grn"

            self.log.tag_config("info", foreground=LOG_FG)
            self.log.tag_config("ok",   foreground=ACCENT)
            self.log.tag_config("err",  foreground=RED)
            self.log.tag_config("red",  foreground="#f87171")
            self.log.tag_config("amb",  foreground="#fbbf24")
            self.log.tag_config("grn",  foreground=ACCENT)

            self.log.insert("end", f"[{ts}] {msg}\n", tag)
            self.log.see("end")
            self.log.configure(state="disabled")
        self.root.after(0, _w)

    def _iniciar(self):
        if not self.arq_cl.get():
            messagebox.showerror("Erro", "Selecione a planilha Clientes Green.")
            return
        if not self.arq_rec.get():
            messagebox.showerror("Erro", "Selecione a planilha de Recebíveis.")
            return
        self.btn.config(state="disabled", text="⏳  Processando...", bg=SURFACE2, fg=TEXT2)
        self.status_var.set("Processando...")
        self._loading = LoadingScreen(self.root)
        threading.Thread(target=self._run, daemon=True).start()

    def _run(self):
        try:
            self._log("Lendo Clientes Green...")
            self.root.after(0, lambda: self._loading.atualizar("Lendo Clientes Green..."))
            df_cl  = self._ler(self.arq_cl.get())
            self._log("Lendo Recebíveis...")
            self.root.after(0, lambda: self._loading.atualizar("Lendo Recebíveis..."))
            df_rec = self._ler(self.arq_rec.get())

            forn = self.forn.get().strip()
            self._log(f"Fornecedora: {forn or 'TODAS'}")
            self.root.after(0, lambda: self._loading.atualizar("Cruzando planilhas..."))

            df = processar(df_cl, df_rec, forn, self._log)

            pasta = os.path.dirname(self.arq_cl.get())
            ds    = datetime.now().strftime("%d-%m-%Y_%H%M")
            fs    = forn.replace(" ","_").upper() if forn else "TODAS"
            path  = os.path.join(pasta, f"ANALISE_FATURAMENTO_{fs}_{ds}.xlsx")

            self._log("Gerando Excel formatado...")
            self.root.after(0, lambda: self._loading.atualizar("Gerando Excel..."))
            salvar_excel(df, path)

            neg  = len(df[df["Status"]=="NEGATIVO"])
            pos  = len(df[df["Status"]=="POSITIVO"])
            zero = len(df[df["Status"]=="ZERO"])
            nf   = len(df[df["Status"]=="AINDA NAO FATURA"])

            self._log("✅  Análise concluída com sucesso!")
            self._log(f"🔴  Faltantes (NEGATIVO):   {neg:,}")
            self._log(f"🟡  Duplicados (POSITIVO):  {pos:,}")
            self._log(f"🟢  OK (ZERO):              {zero:,}")
            self._log(f"⚪  Ainda não fatura:       {nf:,}")
            self._log(f"📁  Arquivo: {os.path.basename(path)}")

            self.root.after(0, lambda: (
                self._loading.fechar(),
                self.status_var.set("✅ Concluído!"),
                messagebox.showinfo("Concluído!",
                    f"Análise finalizada!\n\n"
                    f"🔴 Faltantes:        {neg:,}\n"
                    f"🟡 Duplicados:       {pos:,}\n"
                    f"🟢 OK:               {zero:,}\n"
                    f"⚪ Ainda não fatura: {nf:,}\n\n"
                    f"Salvo em:\n{path}")
            ))
        except Exception as e:
            self._log(f"❌ Erro: {e}")
            self.root.after(0, lambda: (
                self._loading.fechar(),
                self.status_var.set("❌ Erro na execução"),
                messagebox.showerror("Erro", str(e))
            ))
        finally:
            self.root.after(0, lambda: self.btn.config(
                state="normal", text="▶   Processar Análise",
                bg=ACCENT, fg="#000"))

    def _ler(self, path):
        if path.endswith(".csv"):
            for enc in ["utf-8","latin-1","cp1252"]:
                try: return pd.read_csv(path, dtype=str, encoding=enc)
                except: pass
        return pd.read_excel(path, dtype=str)

# ============================================================
if __name__ == "__main__":
    root = tk.Tk()
    App(root)
    root.mainloop()